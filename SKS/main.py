from tkinter import*
from tkinter import messagebox
import openpyxl
from openpyxl import Workbook 
import pathlib 

#Clear Details:
def clear():
    user_name_var.set("")
    phone_no_var.set("")
    A_no_var.set("")
    Door_no_var.set("")
    streetname_var.set("")
    villagename_var.set("")
    Dname_var.set("")
    D1_var.set("")
    D2_var.set("")
    advance_var.set("")
    things_var.set("")

#Search Details:
def search():
    text= search_entry.get()
    clear()
    Upload_button.config(state="disabled")

    file = openpyxl.load_workbook("customer_details.xlsx")
    sheet = file.active

    for row in sheet.iter_rows(min_row=2,max_col=1,max_row=sheet.max_row):
        if row[0].value ==(text):
            name = row[0]
            name_position=str(name)[14:-1]
            Name=str(name)[15:-1]       
    try:
        print(name)
    except:
        messagebox.showerror("Invalid","Customer  Not Found")
    
    X1=sheet.cell(row=int(Name),column=1).value
    X2=sheet.cell(row=int(Name),column=2).value
    X3=sheet.cell(row=int(Name),column=3).value
    X4=sheet.cell(row=int(Name),column=4).value
    X5=sheet.cell(row=int(Name),column=5).value
    X6=sheet.cell(row=int(Name),column=6).value
    X7=sheet.cell(row=int(Name),column=7).value
    X8=sheet.cell(row=int(Name),column=8).value
    X9=sheet.cell(row=int(Name),column=9).value
    X10=sheet.cell(row=int(Name),column=10).value
    X11=sheet.cell(row=int(Name),column=11).value
    
    user_name_var.set(X1)
    phone_no_var.set(X2)
    A_no_var.set(X3)
    Door_no_var.set(X4)
    streetname_var.set(X5)
    villagename_var.set(X6)
    Dname_var.set(X7)
    D1_var.set(X8)
    D2_var.set(X9)
    advance_var.set(X10)
    things_var.set(X11)


# Exit Program:
def exit_program():
    root.destroy()


# Upload Details:
def upload():
    N1 = user_name_var.get()
    P1 = phone_no_var.get()
    A1 = A_no_var.get()
    D1 = Door_no_var.get()
    S1 = streetname_var.get()
    V1 = villagename_var.get()
    D2 = Dname_var.get()
    Date_Hire = D1_var.get()
    Date_Return = D2_var.get()
    A2 = advance_var.get()
    T1 = things_var.get()

    if N1==" " or P1=="" or A1=="" or D1=="" or S1=="" or V1=="" or D2=="" or Date_Hire=="" or Date_Return==""or A2==""or T1=="":
        messagebox.showerror("Error","Some Details Missing")

    else:
        file = openpyxl.load_workbook("customer_details.xlsx")
        sheet=file.active
        sheet.cell(column=1,row=sheet.max_row+1,value=N1)
        sheet.cell(column=2,row=sheet.max_row,value=P1)
        sheet.cell(column=3,row=sheet.max_row,value=A1)
        sheet.cell(column=4,row=sheet.max_row,value=D1)
        sheet.cell(column=5,row=sheet.max_row,value=S1)
        sheet.cell(column=6,row=sheet.max_row,value=V1)
        sheet.cell(column=7,row=sheet.max_row,value=D2)
        sheet.cell(column=8,row=sheet.max_row,value=Date_Hire)
        sheet.cell(column=9,row=sheet.max_row,value=Date_Return)
        sheet.cell(column=10,row=sheet.max_row,value=A2)
        sheet.cell(column=11,row=sheet.max_row,value=T1)
 
        file.save(r"customer_details.xlsx")

        try:
            messagebox.showinfo("Success","The Details uploaded Successfully")
            
        except:
            pass

root=Tk()
root.title("SAMPLE PAGE")
root.geometry("1300x715")
root.resizable(False,False)

file=pathlib.Path("customer_details.xlsx")
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet["A1"]="User Name"
    sheet["B1"]="Phone No"
    sheet["C1"]="Adha No"
    sheet["D1"]="Door No"
    sheet["E1"]="Street Name"
    sheet["F1"]="Village Name"
    sheet["G1"]="District Name"
    sheet["H1"]="Date for Hire"
    sheet["I1"]="Date for Return"
    sheet["J1"]="Advance Amount"
    sheet["K1"]="Things"
    file.save("customer_details.xlsx")

user_name_var = StringVar()
phone_no_var = StringVar()
A_no_var = StringVar()
Door_no_var = StringVar()
streetname_var = StringVar()
villagename_var = StringVar()
Dname_var = StringVar()
D1_var = StringVar()
D2_var = StringVar()
advance_var = StringVar()
things_var = StringVar()

# Icon image:
icon=PhotoImage(file="/home/senkathir/Desktop/SKS/images/pngwing.com (1).png")
root.iconphoto(False,icon)

# Background image:
bg_image=PhotoImage(file="/home/senkathir/Desktop/SKS/images/1300x715.png")
Label(root,image=bg_image).pack()

# Searchbare:
search_entry=Entry(root,width=25,bg="white",fg="black",font=("Arial Bold",20))
search_entry.place(x=786,y=17)

#user name Entry:

user_name=Entry(root,width=38,textvariable=user_name_var,bg="white",fg="black",font=("Arial Bold",20))
user_name.place(x=16,y=182)

# Phone number Entry:    
phone_no_entry=Entry(root,width=38,textvariable=phone_no_var,bg="white",fg="black",font=("Arial Bold",20))
phone_no_entry.place(x=16,y=272)


# Adhar number Entry:
A_no_entry=Entry(root,width=38,textvariable=A_no_var,bg="white",fg="black",font=("Arial Bold",20))
A_no_entry.place(x=16,y=358)

# Door Number Entry:
Door_no_entry=Entry(root,textvariable=Door_no_var,width=30,bg="white",fg="black",font=("Arial Bold",16))
Door_no_entry.place(x=230,y=440)

# Street name Entry:
streetname_entry=Entry(root,textvariable=streetname_var,width=30,bg="white",fg="black",font=("Arial Bold",16))
streetname_entry.place(x=230,y=482)

# Village name Entry:
villagename_entry=Entry(root,textvariable=villagename_var,width=30,bg="white",fg="black",font=("Arial Bold",16))
villagename_entry.place(x=230,y=525)

# Distrit name Entry:
Dname_entry=Entry(root, textvariable=Dname_var,width=30,bg="white",fg="black",font=("Arial Bold",16))
Dname_entry.place(x=230,y=567)

# Date of hire Entry:
D1_entry=Entry(root, textvariable=D1_var,width=38,bg="white",fg="black",font=("Arial Bold",20))
D1_entry.place(x=705,y=182)

# Return date Entry:
D2_entry=Entry(root,textvariable=D2_var,width=38,bg="white",fg="black",font=("Arial Bold",20))
D2_entry.place(x=705,y=272)

# Advance Amount Entry:
advance_entry=Entry(root,textvariable=advance_var,width=38,bg="white",fg="black",font=("Arial Bold",20))
advance_entry.place(x=705,y=358)

# Things Entry:
things_entry=Entry(root, textvariable=things_var,width=38,bg="white",fg="black",font=("Arial Bold",20))
things_entry.place(x=705,y=450)

#Button:
Upload_button=Button(text="UPLOAD",width=14,bd=0,fg="white",bg="black",font=("Arial Bold",16),command=upload)
Upload_button.place(x=573,y=638)

reset_button=Button(text="RESET",width=10,bd=0,fg="white",bg="black",font=("Arial Bold",16),command=clear)
reset_button.place(x=410,y=638)

exit_button=Button(text="EXIT",width=10,bd=0,fg="white",bg="black",font=("Arial Bold",16),command=exit_program)
exit_button.place(x=780,y=638)

search_button=Button(text="SEARCH",width=5,bd=0,fg="black",bg="white",font=("Arial Bold",17),command=search)
search_button.place(x=1174,y=17)

root.mainloop()